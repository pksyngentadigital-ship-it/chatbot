"""Workbook -> Pinecone.

Reads .xlsx with openpyxl rather than pandas. pandas was ~63MB of the
bundle and was used here for little more than "give me the rows and
forward-fill a column", both of which are a few lines of plain Python.

This module does not run on Vercel — ingestion is an admin action driven
from the Streamlit app or a local script, so the serverless function
never has to hold a workbook in memory or survive a long upload.
"""

import hashlib
import re
from io import BytesIO

from vog import parsing as P
from vog import retrieval
from vog.catalog import (
    EMBEDDING_DIMENSION, EMPTY_VALUES, INDEX_STATS_ID, MAX_VALUE_CHARS,
    MONTH_ORDER, NEGATIVE_CATEGORIES, POSITIVE_CATEGORIES,
)

EMBED_BATCH = 96
UPSERT_BATCH = 50
_WEEK_COL_RE = re.compile(r'^w(?:ee)?k(?:\s*(?:no\.?|number|#))?$', re.IGNORECASE)


# ──────────────────────────── cell helpers ───────────────────────────

def _norm(value) -> str:
    """Header/label text with runs of whitespace collapsed."""
    if value is None:
        return ""
    return re.sub(r'\s+', ' ', str(value)).strip()


def _blank(value) -> bool:
    return value is None or not str(value).strip()


def _sheet_rows(ws) -> list[list]:
    """Every row as a plain list. Read-only worksheets are generators, and
    both layouts need random access to the same rows."""
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _make_payload(year, month, week_label, category, bullet,
                  sheet_name, src_row, ingest_run, week_num):
    """Embedding text plus Pinecone metadata for one bullet.

    Provenance (sheet / src_row / ingest_run) is stored so a re-ingest can
    delete exactly what it replaces.
    """
    context = (f"Year: {year}. Month: {month}. Week: {week_label}. "
               f"Case Category: {category}. Feedback: {bullet}.")
    metadata = {
        "month": month,
        "year": year,
        "week": week_label,
        "category": category,
        "sentiment": ("positive" if category in POSITIVE_CATEGORIES
                      else "negative" if category in NEGATIVE_CATEGORIES
                      else "neutral"),
        "value": bullet[:MAX_VALUE_CHARS],
        "crop": ",".join(P.extract_crops(bullet)),
        # Catalog matches only — the price list is the product master.
        "products": ",".join(P.extract_product_mentions(bullet)),
        "sheet": sheet_name,
        "src_row": int(src_row),
        "ingest_run": ingest_run,
    }
    if week_num is not None:
        # An integer, so week can be a database-level filter rather than a
        # Python substring test applied after the top_k cut.
        metadata["week_num"] = int(week_num)
    if len(bullet) > MAX_VALUE_CHARS:
        metadata["truncated"] = True
    return context, metadata


def _vector_id(sheet, category, week_label, row, bullet_idx, bullet) -> str:
    """Content-addressed, so re-ingesting an unchanged row is idempotent.

    The original scheme concatenated the two indexes with no separator, so
    row 1/bullet 12 and row 11/bullet 2 both produced "...112" and one
    silently overwrote the other.
    """
    basis = f"{sheet}|{category}|{week_label}|{row}|{bullet_idx}|{bullet}"
    return "v_" + hashlib.sha1(basis.encode("utf-8")).hexdigest()


# ───────────────────────────── the two layouts ───────────────────────

def _collect_layout_a(rows, sheet, year, ingest_run, skip, tally, out):
    """Categories are ROWS, week columns run across the top."""
    headers = [_norm(c) for c in rows[0]]
    dupes = sorted({h for h in headers if h and headers.count(h) > 1})
    if dupes:
        skip(sheet, "duplicate_columns",
             f"Columns collide after whitespace normalization: {dupes}")
        return

    # Resolved against the whole header list, not per-header: the helper
    # prefers an exact 'Category' over a merely 'categ'-containing header,
    # and asking one at a time would throw that preference away.
    cat_col = P.find_category_column(headers)
    if not cat_col:
        return False  # not layout A after all
    cat_idx = headers.index(cat_col)

    week_idxs = [i for i, h in enumerate(headers) if 'week' in h.lower()]
    if not week_idxs:
        skip(sheet, "no_week_columns",
             "A category column exists but no column header contains 'week'.")
        return

    unmapped = set()
    current_category = None
    for r, row in enumerate(rows[1:], start=1):
        raw = _norm(row[cat_idx]) if cat_idx < len(row) else ""
        # Merged category cells give the value only to the first row, so
        # every continuation row came back empty and was dropped.
        if raw:
            mapped = P.normalize_category(raw)
            if mapped:
                current_category = mapped
            else:
                if raw.lower() not in EMPTY_VALUES:
                    unmapped.add(raw)
                current_category = None
        if not current_category:
            continue

        for c in week_idxs:
            text = P._clean_cell_text(row[c]) if c < len(row) else None
            if text is None:
                continue
            bullets = P.split_bullets(text)
            if not bullets:
                continue
            month = P.extract_month_from_col(headers[c])
            if month == "Unknown":
                skip(sheet, "unparseable_month",
                     f"Week column '{headers[c]}' contains no month name; its rows "
                     f"are unreachable by any month filter and were skipped.")
                continue
            tally(f"{month} {year}", len(bullets))
            for b_idx, bullet in enumerate(bullets):
                out(_make_payload(year, month, headers[c], current_category, bullet,
                                  sheet, r - 1, ingest_run, P._week_number(headers[c])),
                    _vector_id(sheet, current_category, headers[c], r - 1, b_idx, bullet))

    if unmapped:
        skip(sheet, "unmapped_categories",
             f"These category values are not recognized and their rows were "
             f"skipped: {sorted(unmapped)}")
    return True


def _collect_layout_b(rows, sheet, year, ingest_run, skip, tally, out):
    """Month/Week are ROW values, categories are COLUMN headers."""
    header_row = next(
        (i for i in range(min(10, len(rows)))
         if any(_norm(v).lower() == 'month' for v in rows[i])), None)
    if header_row is None:
        skip(sheet, "no_header_row",
             "No row in the first 10 contains a 'Month' header cell.")
        return

    # Forward-fill the header row, so a merged header spanning two columns
    # doesn't silently discard the second column's data.
    col_map, last = {}, None
    for j, v in enumerate(rows[header_row]):
        text = _norm(v)
        if text and text.lower() != 'none':
            last = text
            col_map[j] = text
        elif last is not None:
            col_map[j] = last

    month_idx = next((i for i, v in col_map.items() if v.lower() == 'month'), None)
    week_idx = next((i for i, v in col_map.items() if _WEEK_COL_RE.match(v)), None)

    # Only headers that resolve to a KNOWN category are feedback. Every
    # non-Month/Week column used to qualify, so Region / Remarks / Dealer
    # values were ingested as feedback and region names counted as products.
    category_cols, rejected = {}, []
    for i, v in col_map.items():
        if i in (month_idx, week_idx):
            continue
        if P.normalize_category(v):
            category_cols[i] = v
        else:
            rejected.append(v)

    if month_idx is None or not category_cols:
        skip(sheet, "no_category_columns",
             f"No column header maps to a known feedback category. "
             f"Headers seen: {sorted(set(col_map.values()))}")
        return
    if rejected:
        skip(sheet, "ignored_columns",
             f"Not feedback categories, so not ingested: {sorted(set(rejected))}")

    current_month = current_week = None
    for r in range(header_row + 1, len(rows)):
        row = rows[r]
        mval = row[month_idx] if month_idx < len(row) else None
        if not _blank(mval):
            parsed = P.extract_month_from_col(str(mval).strip())
            if parsed == "Unknown":
                # A 'Total'/'Notes' row is not feedback, and must not
                # overwrite the forward-filled month for the rows beneath.
                continue
            if parsed != current_month:
                # Reset the week too, or a blank week cell on a month's
                # first row inherits the previous month's last week.
                current_week = None
            current_month = parsed

        if week_idx is not None and week_idx < len(row) and not _blank(row[week_idx]):
            current_week = str(row[week_idx]).strip()

        if not current_month:
            continue
        week_label = f"{current_week} Week {current_month}" if current_week else current_month

        for c, header in category_cols.items():
            category = P.normalize_category(header)
            text = P._clean_cell_text(row[c]) if c < len(row) else None
            if not category or text is None:
                continue
            bullets = P.split_bullets(text)
            if not bullets:
                continue
            tally(f"{current_month} {year}", len(bullets))
            for b_idx, bullet in enumerate(bullets):
                out(_make_payload(year, current_month, week_label, category, bullet,
                                  sheet, r, ingest_run, P._week_number(week_label)),
                    _vector_id(sheet, category, week_label, r, b_idx, bullet))


# ────────────────────────────── entry point ──────────────────────────

def run_ingestion(file_bytes: bytes, pinecone_api_key: str | None = None,
                  purge_first: bool = False, dry_run: bool = False) -> dict:
    """Parse, tag, embed and upsert. Returns total_records / summary /
    skipped / ingest_run.

    `skipped` is load-bearing: four silent `continue` paths could once drop
    entire sheets while the call still reported success, so a workbook
    could half-ingest with no indication anything was wrong.

    purge_first deletes the index before writing. Needed after any change
    to the tagging logic, because vectors already in the index keep their
    old metadata — re-ingesting alone does not repair them.

    dry_run parses and tags everything but writes nothing and needs no API
    key. The whole parse happens before the first network call, so this is
    the real result — you can check a workbook against production's exact
    rules without touching production.
    """
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    sheet_names = list(wb.sheetnames)
    ingest_run = hashlib.sha1(
        f"{len(file_bytes)}|{','.join(sheet_names)}".encode()).hexdigest()[:12]

    payloads, texts, summary, skipped = [], [], {}, []

    def skip(sheet, reason, detail=""):
        skipped.append({"sheet": str(sheet), "reason": reason, "detail": str(detail)})

    def tally(key, n):
        summary[key] = summary.get(key, 0) + n

    def out(payload_and_text, vec_id):
        context, metadata = payload_and_text
        payloads.append({"id": vec_id, "metadata": metadata})
        texts.append(context)

    for name in sheet_names:
        sheet = str(name).strip()
        year = P.infer_year_for_sheet(sheet, [str(s) for s in sheet_names])
        if not year:
            skip(sheet, "no_year",
                 "Could not infer a year for this sheet, or the inferred year "
                 "would duplicate an explicitly dated sheet.")
            continue

        rows = _sheet_rows(wb[name])
        if not rows:
            skip(sheet, "empty_sheet", "The sheet has no rows.")
            continue

        if _collect_layout_a(rows, sheet, year, ingest_run, skip, tally, out) is False:
            _collect_layout_b(rows, sheet, year, ingest_run, skip, tally, out)

    wb.close()

    # Collapse duplicate ids before embedding, so the reported count matches
    # what actually lands in the index.
    seen, deduped, deduped_texts = set(), [], []
    for item, text in zip(payloads, texts):
        if item["id"] in seen:
            continue
        seen.add(item["id"])
        deduped.append(item)
        deduped_texts.append(text)
    payloads, texts = deduped, deduped_texts

    if not payloads:
        raise ValueError("No records found in the uploaded workbook."
                         + (f" Skipped: {skipped}" if skipped else ""))

    if dry_run:
        return {"total_records": len(payloads), "summary": summary,
                "skipped": skipped, "ingest_run": ingest_run, "dry_run": True}

    if not pinecone_api_key:
        raise ValueError("A Pinecone API key is required to write. "
                         "Pass dry_run=True to parse without writing.")

    pc, index = retrieval.connect(pinecone_api_key)

    if purge_first:
        try:
            index.delete(delete_all=True)
        except Exception as e:
            raise RuntimeError(
                f"Purge requested but the index delete failed, so ingestion was "
                f"aborted to avoid mixing old and new data: {e}")

    # Embed and upsert in the same pass. Collecting every embedding first
    # held the whole dataset in memory (~12KB/record), which OOMs a small
    # instance well before a realistic workbook finishes.
    written = 0
    for i in range(0, len(payloads), EMBED_BATCH):
        meta_batch = payloads[i:i + EMBED_BATCH]
        resp = pc.inference.embed(
            model="llama-text-embed-v2",
            inputs=texts[i:i + EMBED_BATCH],
            parameters={"input_type": "passage", "dimension": EMBEDDING_DIMENSION},
        )
        values = [item.values for item in resp]
        if len(values) != len(meta_batch):
            # Positional pairing means a short response would silently
            # attach every later vector to the wrong metadata.
            raise RuntimeError(
                f"Embedding count mismatch ({len(values)} vectors for "
                f"{len(meta_batch)} records); aborted after {written} records "
                f"to avoid writing mismatched data.")
        vectors = [{"id": m["id"], "values": v, "metadata": m["metadata"]}
                   for m, v in zip(meta_batch, values)]
        for j in range(0, len(vectors), UPSERT_BATCH):
            index.upsert(vectors=vectors[j:j + UPSERT_BATCH])
        written += len(vectors)

    _write_stats(index, payloads, written, ingest_run)

    return {"total_records": written, "summary": summary,
            "skipped": skipped, "ingest_run": ingest_run}


def _write_stats(index, payloads, written, ingest_run) -> None:
    """Record the dataset's real extents, so relative-date resolution reads
    a fact instead of sampling arbitrary records and taking a max."""
    dated = [(int(m["metadata"]["year"]), m["metadata"]["month"]) for m in payloads
             if str(m["metadata"].get("year", "")).isdigit()
             and m["metadata"].get("month") in MONTH_ORDER]
    if not dated:
        return
    max_year, max_month = max(dated, key=lambda p: (p[0], MONTH_ORDER[p[1]]))
    min_year, min_month = min(dated, key=lambda p: (p[0], MONTH_ORDER[p[1]]))
    try:
        index.upsert(vectors=[{
            "id": INDEX_STATS_ID,
            "values": [0.0] * EMBEDDING_DIMENSION,
            "metadata": {
                "is_stats_record": True,
                "max_year": str(max_year), "max_month": max_month,
                "min_year": str(min_year), "min_month": min_month,
                "total_records": written, "ingest_run": ingest_run,
            },
        }])
    except Exception:
        pass  # an optimization; the sampled fallback still works
